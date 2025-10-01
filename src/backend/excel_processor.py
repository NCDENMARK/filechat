import os
import hashlib
from pathlib import Path
from typing import List, Dict
import openpyxl

class ExcelProcessor:
    def __init__(self):
        # Keep track of files we've already processed
        self.processed_files = set()
    
    def scan_directory(self, directory_path: str) -> List[str]:
        """Find all Excel files in a given folder"""
        excel_files = []  # List to store found Excel paths
        path = Path(directory_path)  # Convert string to Path object
        
        # Check if the directory exists and is actually a directory
        if path.exists() and path.is_dir():
            # Find all .xlsx files (case-insensitive)
            for excel_path in path.glob("*.[xX][lL][sS][xX]"):
                excel_files.append(str(excel_path))  # Add full path to list
            # Also check for older .xls files
            for excel_path in path.glob("*.[xX][lL][sS]"):
                # Only add if it's not already in the list (avoid duplicates)
                path_str = str(excel_path)
                if not path_str.endswith('.xlsx') and not path_str.endswith('.XLSX'):
                    excel_files.append(path_str)
        
        return excel_files  # Return list of Excel file paths
    
    def extract_text(self, excel_path: str) -> Dict:
        """Pull all data out of an Excel file, converting it to readable text"""
        try:
            # Open the Excel workbook
            workbook = openpyxl.load_workbook(excel_path, data_only=True)
            text = ""  # String to accumulate all text
            
            # Process each sheet in the workbook
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text += f"\n\n=== Sheet: {sheet_name} ===\n\n"
                
                # Get the data from the sheet
                rows = list(sheet.values)
                if not rows:
                    text += "[Empty sheet]\n"
                    continue
                
                # Assume first row is headers (common in Excel files)
                headers = rows[0] if rows else []
                
                # Convert headers to strings, handling None values
                headers = [str(h) if h is not None else f"Column{i+1}" 
                          for i, h in enumerate(headers)]
                
                # Add header row
                text += "Headers: " + " | ".join(headers) + "\n\n"
                
                # Process each data row
                for row_num, row in enumerate(rows[1:], start=2):
                    row_data = []
                    for col_num, (header, cell_value) in enumerate(zip(headers, row)):
                        if cell_value is not None and str(cell_value).strip():
                            # Create readable key-value pairs
                            row_data.append(f"{header}: {cell_value}")
                    
                    if row_data:
                        text += f"Row {row_num}: " + ", ".join(row_data) + "\n"
                
                text += "\n"  # Add spacing between sheets
            
            # Create a unique ID for this file using MD5 hash
            file_id = hashlib.md5(excel_path.encode()).hexdigest()
            
            # Count sheets
            sheet_count = len(workbook.sheetnames)
            
            # Return all the extracted information
            return {
                "file_id": file_id,  # Unique identifier
                "file_path": excel_path,  # Full path to file
                "file_name": Path(excel_path).name,  # Just the filename
                "text": text,  # All extracted text
                "sheet_count": sheet_count,  # Number of sheets
                "status": "success"  # Extraction worked
            }
        except Exception as e:
            # If something goes wrong, return error info
            return {
                "file_path": excel_path,
                "file_name": Path(excel_path).name,
                "status": "error",
                "error": str(e)  # What went wrong
            }
    
    def split_into_chunks(self, text: str, chunk_size: int = 500, overlap: int = 50) -> List[str]:
        """Break long text into smaller, overlapping pieces"""
        words = text.split()  # Split text into individual words
        chunks = []  # List to store text chunks
        
        # Move through the text in steps of (chunk_size - overlap)
        for i in range(0, len(words), chunk_size - overlap):
            # Take chunk_size words starting from position i
            chunk = ' '.join(words[i:i + chunk_size])
            if chunk:  # Only add non-empty chunks
                chunks.append(chunk)
        
        return chunks